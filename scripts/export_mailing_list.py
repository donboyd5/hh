"""Build and export the potential-donor mailing list (local-only, PII).

Reads the latest Neon pull (offline; no API needed), the processed registration table, and
the external workbooks, then writes:

  data/20_processed/mailing_list.parquet
  data/20_processed/hh-mailing-list.xlsx   (mailing-list sheet + about sheet)

Everything names real people: the outputs stay under gitignored ``data/`` and must never be
published or committed. Rerun after ``scripts/pull.py`` + ``scripts/build.py`` and after any
edit to the external workbooks — the build is deterministic given the same inputs.

Usage:
    python scripts/export_mailing_list.py
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from hh import config, io
from hh.analytics.fst_match import fuzzy_fst_candidates
from hh.analytics.mailing import build_mailing_list
from hh.clean.accounts import clean_accounts
from hh.clean.donations import clean_donations
from hh.external import fortsalem as fst
from hh.external import mailing as ml
from hh.external.fst_confirmations import load_confirmations, resolve
from hh.external.mailing import match_households
from hh.external.notes import load_boyd_notes, load_fst_contact_notes, load_fst_web_notes
from hh.external.provenance import append_external_manifest, external_source_entry

XLSX_FILENAME = "hh-mailing-list.xlsx"
JUDY_FILENAME = "fst-donors-in-neon.xlsx"

# Hand-maintained aliases for workbook names that no longer match Neon exactly (Neon
# household names get edited between pulls; the workbooks age). Keyed by the workbook's
# name -> rollup id. Add entries when the UNMATCHED list in the export output names a
# household you know is in Neon under a different name.
HOUSEHOLD_ALIASES = {
    # Neon renamed this household between the 2026-07-07 and 2026-08-27 pulls
    # (account 39722, household 1684); Don's silent-1000plus note must still attach
    "Tim Smith & Elizabeth Straiton": "1684",
}


def _matched(frame: pd.DataFrame, households: pd.DataFrame) -> pd.DataFrame:
    """A household_name-keyed external frame with the matched rollup ``id`` attached.

    Uses the city tiebreak when the frame carries a city (Judy's exports do), then the
    hand-maintained aliases for anything still unmatched.
    """
    cities = frame["city"] if "city" in frame.columns else None
    matched = match_households(frame["household_name"], households, cities=cities)
    ids = list(matched["id"].values)
    how = list(matched["match"].values)
    for i, name in enumerate(frame["household_name"]):
        if pd.isna(ids[i]) and str(name).strip() in HOUSEHOLD_ALIASES:
            ids[i] = HOUSEHOLD_ALIASES[str(name).strip()]
            how[i] = "alias"
    return frame.assign(id=ids, match=how)


def _union_years(values) -> str:
    """Sorted union of comma-separated year lists ("2021,2024" + "2024,2025")."""
    return ",".join(sorted({y for v in values for y in str(v).split(",") if y}))


def _freeze_header(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _format_review_sheet(ws, *, n_rows: int) -> None:
    """Make the review sheet obvious to mark: wide bold CONFIRM column with a Y/N
    dropdown, frozen header, readable widths on the name columns."""
    _freeze_header(ws)
    widths = {"confirm": 10, "boyd_note": 30, "status": 22, "fst_name": 34,
              "fst_best_tier": 20, "matched_via": 30, "neon_household": 34,
              "neon_city": 16, "web_note": 60}
    for cell in ws[1]:
        if cell.value in widths:
            ws.column_dimensions[cell.column_letter].width = widths[cell.value]
    ws["A1"].fill = PatternFill("solid", fgColor="FFF2CC")
    if n_rows:
        dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
        dv.prompt, dv.promptTitle = "Y = same household, N = different", "Confirm"
        ws.add_data_validation(dv)
        dv.add(f"A2:A{n_rows + 1}")
        for r in range(2, n_rows + 2):
            ws[f"A{r}"].fill = PatternFill("solid", fgColor="FFF2CC")


def _fold_fst_contacts(table: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Fill address fields on Fort Salem (needs_review) rows from the research outputs.

    Returns the table and the research sheet (every roll hit plus web notes, so Don can
    see the runner-up address and the evidence). Deaths confirmed on the web drop the row
    (printed); a named survivor re-labels the household.
    """
    contacts_path = config.layer_dir("interim") / "fst_contacts.parquet"
    hits = pd.read_parquet(contacts_path) if contacts_path.exists() else pd.DataFrame(
        columns=["household_name", "given_agreement", "owners", "street", "city", "state",
                 "zip", "town", "source", "confidence"]
    )
    notes = load_fst_contact_notes()
    attrs = dict(table.attrs)

    dead = set(notes.loc[notes["deceased"], "household_name"])
    is_dead = table["needs_review"] & table["household_name"].isin(dead)
    dropped = table.loc[is_dead, "household_name"]
    for name in dropped:
        print(f"  dropped (web-confirmed deceased): {name}")
    table = table[~(table["needs_review"] & table["household_name"].isin(dead))].copy()
    survivors = dict(zip(notes["household_name"], notes["survivor"], strict=True))
    relabel = table["needs_review"] & table["household_name"].map(survivors).notna()
    table.loc[relabel, "household_name"] = table.loc[relabel, "household_name"].map(survivors)

    best = hits.drop_duplicates("household_name").set_index("household_name")
    fst = table["needs_review"]
    field_map = [("address", "street"), ("city", "city"), ("state_province", "state"), ("zip_code", "zip")]
    for col, src in field_map:
        table.loc[fst, col] = table.loc[fst, "household_name"].map(best[src])
    table["address_source"] = pd.NA
    table.loc[fst, "address_source"] = table.loc[fst, "household_name"].map(
        best["source"] + " (" + best["confidence"] + "; owner: " + best["owners"].astype(str) + ")"
    )
    note_map = notes.set_index("household_name")["contact_note"]
    table["contact_note"] = pd.NA
    table.loc[fst, "contact_note"] = table.loc[fst, "household_name"].map(note_map)
    table.attrs.update(attrs)

    research = table.loc[fst, ["household_name", "fst_best_tier", "fst_years_list"]].merge(
        hits, on="household_name", how="left"
    ).merge(notes.drop(columns=["deceased", "survivor"]), on="household_name", how="left")
    return table, research.sort_values(["household_name", "given_agreement"], ascending=[True, False])


def main() -> None:
    # processed tables carry the geocode/distance and event-category enrichment (build.py);
    # donations are re-cleaned from the latest pull since no processed donations table exists
    accounts = io.read_parquet("processed", "accounts_geocoded.parquet")
    donations = clean_donations(accounts=clean_accounts())
    registrations = io.read_parquet("processed", "registrations_enriched.parquet")
    appeal_hh = io.read_parquet("processed", "appeal_households.parquet")  # appealed flag
    households = accounts.drop_duplicates(subset=["id"])[["id", "name", "city"]]

    donor3 = _matched(ml.load_donor3(), households)
    # only new accounts with >= $100 lifetime registrations (Judy's column) qualify
    new_accounts_all = ml.load_new_accounts()
    new_accounts = _matched(ml.qualifying_new_accounts(new_accounts_all), households)
    djb = ml.load_djb_workbook()
    silent_selected = _matched(djb["silent_selected"], households)
    appeal_responded = _matched(djb["appeal_responded"], households)

    # Fort Salem: individuals only (the spec excludes orgs and anonymous listings)
    sponsors = fst.load_fst_sponsors()
    fst_summary = fst.summarize_sponsors(sponsors)
    fst_summary = fst_summary[~fst_summary["anonymous"] & ~fst_summary["org"]]
    fst_match = match_households(fst_summary["name"], households)
    fst_summary = fst_summary.assign(
        id=fst_match["id"].values,
        in_neon=fst_match["match"].isin(["name+city", "name"]),
        match_type=fst_match["match"].map({"name+city": "exact", "name": "exact"}),
    )

    # Don's confirmed fuzzy matches fold into their Neon households (fst-confirmations.yaml)
    decisions = load_confirmations()
    confirmed, duplicates, conflicts = resolve(decisions, households)
    folded = fst_summary["name"].isin(confirmed) & ~fst_summary["in_neon"]
    fst_summary.loc[folded, "id"] = fst_summary.loc[folded, "name"].map(confirmed)
    fst_summary.loc[folded, "in_neon"] = True
    fst_summary.loc[folded, "match_type"] = "confirmed"
    boyd_notes = load_boyd_notes()
    for r in decisions[(decisions["confirm"] == "Y") & decisions["boyd_note"].notna()].itertuples():
        hh_id = confirmed.get(r.fst_name)
        if hh_id and r.boyd_note:
            prior = boyd_notes.get(hh_id)
            boyd_notes[hh_id] = f"{prior}; {r.boyd_note}" if prior else str(r.boyd_note)

    table = build_mailing_list(
        accounts,
        donations,
        registrations,
        donor3=donor3,
        new_accounts=new_accounts,
        silent_selected=silent_selected,
        appeal_responded=appeal_responded,
        fst_summary=fst_summary,
        boyd_notes=boyd_notes,
        appealed_ids=set(appeal_hh.loc[appeal_hh["appealed"], "id"].astype(str)),
    )

    # Fort Salem fuzzy candidates for Don to confirm (review sheet; never auto-merged)
    # review sheet: candidates for names still unmatched, plus every decided pair (so
    # Don sees his marks pre-filled and can change them; harvest picks changes up)
    pending = fst_summary[fst_summary["match_type"].ne("confirmed")]
    fst_review = fuzzy_fst_candidates(pending, accounts)
    decided_pairs = decisions.rename(columns={"confirm": "confirm", "boyd_note": "boyd_note"})
    fst_review = fst_review.merge(
        decided_pairs[["fst_name", "neon_hh_id", "confirm", "boyd_note"]],
        on=["fst_name", "neon_hh_id"], how="outer",
    )
    fst_review["status"] = np.select(
        [fst_review["fst_name"].isin(conflicts), fst_review["confirm"].eq("Y"),
         fst_review["confirm"].eq("N")],
        ["CONFLICT - two different households confirmed", "confirmed", "rejected"],
        default="undecided",
    )
    fst_review["web_note"] = fst_review["fst_name"].map(load_fst_web_notes())
    fst_review = fst_review.sort_values(["fst_name", "rank"], na_position="last")
    fst_review = fst_review[
        ["confirm", "boyd_note", "status", "fst_name", "fst_best_tier", "fst_years", "rank",
         "score", "surname_ratio", "given_agreement", "matched_via", "neon_hh_id",
         "neon_household", "neon_city", "web_note"]
    ].reset_index(drop=True)

    # Fort Salem donors who ARE in Neon (exact or confirmed): the record for Judy to add
    # a Neon field from. One row per Neon household; duplicate person records listed.
    in_neon = fst_summary[fst_summary["in_neon"]].copy()
    judy = (
        in_neon.groupby("id")
        .agg(
            fst_names=("name", lambda s: "; ".join(sorted(set(s)))),
            fst_best_tier=("best_tier", "first"),
            fst_years=("years", _union_years),
            match_type=("match_type", "first"),
        )
        .reset_index()
        .rename(columns={"id": "neon_hh_id"})
    )
    judy = judy.merge(
        table[["neon_hh_id", "neon_account_ids", "household_name", "city"]], on="neon_hh_id", how="left"
    )
    missing_hh = judy["household_name"].isna()  # households not in the mailing list
    judy.loc[missing_hh, "household_name"] = judy.loc[missing_hh, "neon_hh_id"].map(
        dict(zip(households["id"].astype(str), households["name"], strict=True))
    )
    judy.loc[missing_hh, "neon_account_ids"] = judy.loc[missing_hh, "neon_hh_id"].map(
        accounts.groupby("id")["account_id"].agg(lambda s: ",".join(sorted(s.astype(str))))
    )
    judy["duplicate_neon_ids"] = judy["neon_hh_id"].map(
        {k: ",".join(v) for k, v in duplicates.items()}
    )
    judy["in_mailing_list"] = ~missing_hh
    judy["neon_field_value"] = "FST donor " + judy["fst_years"]
    judy["boyd_note"] = judy["neon_hh_id"].map(boyd_notes)  # incl. notes from the review sheet
    judy = judy[
        ["neon_hh_id", "neon_account_ids", "duplicate_neon_ids", "household_name", "city",
         "fst_names", "fst_best_tier", "fst_years", "match_type", "in_mailing_list",
         "neon_field_value", "boyd_note"]
    ].sort_values("household_name").reset_index(drop=True)
    io.write_parquet(judy, "processed", "fst_donors_in_neon.parquet")
    with pd.ExcelWriter(config.layer_dir("processed") / JUDY_FILENAME, engine="openpyxl") as xw:
        judy.to_excel(xw, sheet_name="fst-donors-in-neon", index=False)
        _freeze_header(xw.sheets["fst-donors-in-neon"])

    # Fort Salem people not in Neon: addresses from the assessment-roll research
    # (scripts/research_fst_contacts.py) and web notes; web-confirmed deaths drop
    table, fst_research = _fold_fst_contacts(table)

    qa = table.attrs.get("exclusion_qa", {})
    io.write_parquet(table, "processed", "mailing_list.parquet")
    io.write_parquet(fst_review, "processed", "fst_candidates.parquet")
    xlsx_path = config.layer_dir("processed") / XLSX_FILENAME
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as xw:
        table.to_excel(xw, sheet_name="mailing-list", index=False)
        fst_review.to_excel(xw, sheet_name="fst-candidates", index=False)
        fst_research.to_excel(xw, sheet_name="fst-contact-research", index=False)
        _freeze_header(xw.sheets["fst-contact-research"])
        pd.DataFrame({"household_name": qa.get("do_not_contact", [])}).to_excel(
            xw, sheet_name="do-not-contact", index=False
        )
        fst_dropped = pd.DataFrame(table.attrs.get("fst_dropped", []))
        fst_dropped.to_excel(xw, sheet_name="fst-dropped", index=False)
        _freeze_header(xw.sheets["fst-dropped"])
        _format_review_sheet(xw.sheets["fst-candidates"], n_rows=len(fst_review))
        _freeze_header(xw.sheets["mailing-list"])
        pd.DataFrame(
            {
                "item": [
                    "rows", "in_neon", "needs_review (Fort Salem, not in Neon)",
                    "src_donor_5yr", "src_donor3", "src_new_accounts",
                    "letter = donor", "letter = class-family", "letter = new-attender",
                    "letter = fst-personal (not in the appeal)",
                    "src_silent_selected", "src_appeal_responded",
                    "src_appeal_gift (>= $10 in Oct 2025-Jan 2026)",
                    "src_engaged_nondonor (>= $500 FY22-26 spend, no gift)", "fst flagged",
                    "do_not_contact (flagged, kept; see sheet)", "deceased",
                ],
                "detail": [
                    len(table),
                    int(table["in_neon"].sum()),
                    int(table["needs_review"].sum()),
                    int(table["src_donor_5yr"].sum()),
                    int(table["src_donor3"].sum()),
                    int(table["src_new_accounts"].sum()),
                    int(table["letter"].eq("donor").sum()),
                    int(table["letter"].eq("class-family").sum()),
                    int(table["letter"].eq("new-attender").sum()),
                    int(table["letter"].eq("fst-personal").sum()),
                    int(table["src_silent_selected"].sum()),
                    int(table["src_appeal_responded"].sum()),
                    int(table["src_appeal_gift"].sum()),
                    int(table["src_engaged_nondonor"].sum()),
                    int(table["fst"].sum()),
                    len(qa.get("do_not_contact", [])),
                    int(table["deceased"].fillna(False).sum()),
                ],
            }
        ).to_excel(xw, sheet_name="about", index=False)

    # provenance: one entry per distinct file version of each hand-maintained source
    for filename, note in [
        (ml.DONOR3_FILENAME, "Judy's FY24-FY26 donor export with Don's Ambassador/notes"),
        (ml.NEW_ACCOUNTS_FILENAME, "Judy's FY25-26 new-accounts export for AF mailing"),
        (
            ml.DJB_WORKBOOK_FILENAME,
            "Don's edited donor workbook (bold silent-1000plus, action notes)",
        ),
    ]:
        path: Path = config.layer_dir("external") / filename
        if path.exists():
            append_external_manifest(external_source_entry(path, note), slug="mailing-list")

    unmatched = {
        source: frame.loc[frame["id"].isna(), "household_name"].tolist()
        for source, frame in [
            ("donor3", donor3),
            ("new_accounts", new_accounts),
            ("silent_selected", silent_selected),
            ("appeal_responded", appeal_responded),
        ]
    }
    qa = table.attrs.get("exclusion_qa", {})
    print(
        f"new accounts: {len(new_accounts)} of {len(new_accounts_all)} clear "
        f"${ml.MIN_NEW_ACCOUNT_REGISTRATION:.0f} in lifetime registrations"
    )
    print(
        f"fst: {int(fst_summary['in_neon'].sum())} sponsors in Neon "
        f"({int(fst_summary['match_type'].eq('confirmed').sum())} by Don's confirmation, "
        f"{len(judy)} households -> {JUDY_FILENAME}); "
        f"{int((~fst_summary['in_neon']).sum())} not in Neon; "
        f"review sheet: {int(fst_review['status'].eq('undecided').sum())} undecided rows"
    )
    for name in conflicts:
        print(f"  CONFLICT (held, not folded): {name} confirmed against two different households")
    print(
        f"fst rule B: {int(table['needs_review'].sum())} Fort Salem sponsors kept, "
        f"{len(table.attrs.get('fst_dropped', []))} dropped -> fst-dropped sheet"
    )
    print(
        f"mailing list: {len(table)} rows "
        f"({int(table['in_neon'].sum())} in Neon, {int(table['needs_review'].sum())} FST new)"
    )
    print(
        f"exclusions: {len(qa.get('do_not_contact', []))} do-not-contact FLAGGED (kept; see sheet), "
        f"{len(qa.get('dropped_deceased_neon', []))} deceased (Neon flag), "
        f"{len(qa.get('dropped_deceased_note', []))} deceased (notes), "
        f"{len(qa.get('dropped_small_donor', []))} rows under $200 not keep-identified"
    )
    for label in ("dropped_deceased_note", "kept_deceased_note_survivor"):
        names = qa.get(label, [])
        if names:
            print(f"  {label}: {names}")
    for source, names in unmatched.items():
        if names:
            print(f"  UNMATCHED {source} ({len(names)}): {names[:8]}")
    print(f"saved: data/20_processed/mailing_list.parquet and {XLSX_FILENAME}")


if __name__ == "__main__":
    main()
