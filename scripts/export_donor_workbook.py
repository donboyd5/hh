"""Export the donor workbook (Excel) for stewardship use — LOCAL ONLY, PII.

Writes data/20_processed/hh-donor-workbook.xlsx (gitignored) with three sheets:

  donors          every household with pre-appeal individual giving, with appeal
                  response, prior-giving history, engagement, and Boyd notes
  silent-1000plus prior $1,000+ appealed households that did not respond (the
                  analysis-site table, uncapped), with Boyd notes
  about           definitions, sources, and build info

Run after `scripts/build.py`. Notes come from data/30_external/boyd-notes.yaml
(editable; gitignored). Never publish or commit the workbook.
"""
from __future__ import annotations

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from hh import io
from hh.analytics.appeal import last_gift_year_totals
from hh.clean.accounts import clean_accounts
from hh.clean.donations import clean_donations
from hh.external.notes import load_boyd_notes
from hh.provenance import manifest

OUTPUT = "hh-donor-workbook.xlsx"

MONEY_COLS = [
    "lifetime before appeal", "last-gift-yr $", "appeal gifts $", "FY25 giving $",
]


def _town_with_state(df: pd.DataFrame) -> pd.Series:
    """Town with ', ST' appended when out of state; bare state when the town is missing."""

    def join(row) -> object:
        city = "" if pd.isna(row["city"]) else str(row["city"]).strip()
        state = "" if pd.isna(row["state_province"]) else str(row["state_province"]).strip()
        if not city:
            return state if state and state != "NY" else pd.NA
        return f"{city}, {state}" if state and state != "NY" else city

    return df.apply(join, axis=1)


def build_sheets() -> dict[str, pd.DataFrame]:
    t = io.read_parquet("processed", "appeal_households.parquet")
    last_yr = last_gift_year_totals(clean_donations(accounts=clean_accounts()))
    notes = load_boyd_notes()
    t = t.merge(last_yr, on="id", how="left")
    t["town"] = _town_with_state(t)
    t["note"] = t["id"].astype(str).map(notes)

    donors = t[t["prior_donor"]].copy()
    donors = pd.DataFrame(
        {
            "household": donors["name"],
            "town": donors["town"],
            "appealed": donors["appealed"],
            "segment": donors["appeal_segment"],
            "responded": donors["responded"],
            "appeal gifts $": donors["afd_amount"],
            "lifetime before appeal": donors["prior_lifetime_amount"],
            "prior tier": donors["prior_size_tier"],
            "gave FY25": donors["prior_fy_donor"],
            "lapsed 24mo+": donors["lapsed_donor"],
            "last gift": donors["prior_last_gift"],
            "last-gift-yr $": donors["last_gift_year_total"],
            "engagement": donors["eng_profile"],
            "distance": donors["distance_band"],
            "note": donors["note"],
        }
    ).sort_values("lifetime before appeal", ascending=False)

    silent = t[
        t["appealed"] & t["prior_donor"] & ~t["responded"] & (t["prior_lifetime_amount"] >= 1000)
    ].copy()
    silent = pd.DataFrame(
        {
            "household": silent["name"],
            "town": silent["town"],
            "segment": silent["appeal_segment"],
            "lifetime before appeal": silent["prior_lifetime_amount"],
            "last gift": silent["prior_last_gift"],
            "last-gift-yr $": silent["last_gift_year_total"],
            "engagement": silent["eng_profile"],
            "note": silent["note"],
        }
    ).sort_values("lifetime before appeal", ascending=False)

    about = pd.DataFrame(
        {
            "item": [
                "generated",
                "git commit",
                "source table",
                "donors sheet",
                "silent-1000plus sheet",
                "appeal gifts $",
                "lifetime before appeal",
                "last-gift-yr $",
                "engagement",
                "note",
                "PII",
            ],
            "detail": [
                manifest.now_iso(),
                manifest.git_commit(),
                "data/20_processed/appeal_households.parquet (latest build)",
                "all households with individual giving before Oct 1 2025",
                f"appealed prior donors with $1,000+ lifetime giving that did not "
                f"respond (all {len(silent)} of them, not capped at 25)",
                "SUCCEEDED gifts to Annual Fund Drive - 2025-2026, Oct 2025 - Jan 2026",
                "lifetime SUCCEEDED individual gifts before Oct 1 2025",
                "total given in the calendar year of the most recent pre-appeal gift",
                "arts / class / arts+class / community-only / none, as of Oct 1 2025 "
                "(event start dates)",
                "Don's stewardship notes from data/30_external/boyd-notes.yaml (editable)",
                "LOCAL ONLY — contains names and giving; never commit or publish",
            ],
        }
    )
    return {"donors": donors, "silent-1000plus": silent, "about": about}


def write_workbook(sheets: dict[str, pd.DataFrame]) -> None:
    path = io.config.layer_dir("processed") / OUTPUT
    with pd.ExcelWriter(path, engine="openpyxl", datetime_format="yyyy-mm-dd") as xl:
        for sheet, df in sheets.items():
            df.to_excel(xl, sheet_name=sheet, index=False)
            ws = xl.sheets[sheet]
            ws.freeze_panes = "A2"
            for j, col in enumerate(df.columns, start=1):
                letter = get_column_letter(j)
                widest = int(df[col].astype(str).str.len().max() or 12)
                ws.column_dimensions[letter].width = max(12, min(26, widest + 2))
                ws.cell(row=1, column=j).font = Font(bold=True)
                ws.cell(row=1, column=j).alignment = Alignment(vertical="top")
                if col in MONEY_COLS:
                    for i in range(2, len(df) + 2):
                        ws.cell(row=i, column=j).number_format = '$#,##0'
    manifest.append_run_log(
        {
            "recorded_at": manifest.now_iso(),
            "stage": "donor-workbook",
            "git_commit": manifest.git_commit(),
            "rows": {sheet: len(df) for sheet, df in sheets.items()},
            "file": OUTPUT,
        }
    )
    print(f"wrote {path} ({', '.join(f'{s}={len(df)}' for s, df in sheets.items())})", flush=True)


def main() -> None:
    write_workbook(build_sheets())


if __name__ == "__main__":
    main()
