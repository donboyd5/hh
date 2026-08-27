"""Potential-donor mailing-list inputs: three hand-maintained workbooks in data/30_external.

Sources (see meta-docs/create_list.md for the spec):

- ``1.HH donor last 3 years wo Board_djb.xlsx`` — Judy's Neon report of donor households
  FY24–FY26, with Don's hand-added ``Ambassador`` (→ *steward*) and ``notes`` columns.
  Carries no Neon ids; households are matched by name to the rollup.
- ``New Accounts 25-26 for AF Mailing.xlsx`` — Judy's export of accounts new in FY25–FY26
  (header on workbook row 4), with a few hand notes typed into headerless columns at the
  right edge. Also matched by name/email, no Neon ids.
- ``hh-donor-workbook_djb.xlsx`` — Don's edited copy of the generated donor workbook. The
  edits live in formatting and added columns, not new rows: *bold* household names on the
  ``silent-1000plus`` sheet mark the selected non-responders to keep, and an ``action??``
  column holds follow-up notes to merge into the note field.

All three workbooks name people, so they stay under gitignored ``data/`` and nothing from
here is published. The loaders return household *names* plus hand fields; matching to the
rollup ids happens in :func:`match_households` so tests can run on synthetic frames.
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

from .. import config

DONOR3_FILENAME = "1.HH donor last 3 years wo Board_djb.xlsx"
NEW_ACCOUNTS_FILENAME = "New Accounts 25-26 for AF Mailing.xlsx"
DJB_WORKBOOK_FILENAME = "hh-donor-workbook_djb.xlsx"


def _ext_path(filename: str) -> Path:
    return config.layer_dir("external") / filename


# -- name matching --------------------------------------------------------------
def norm_name(series: pd.Series) -> pd.Series:
    """Normalize a household-name column for matching: casefold, collapse internal whitespace.

    Judy's export contains at least one double-spaced name ("andrea  strebel"), so a plain
    lower/strip join would miss real matches.
    """
    return (
        series.astype("string")
        .str.strip()
        .str.lower()
        .str.replace(r"\s+", " ", regex=True)
    )


def match_households(
    names: pd.Series,
    households: pd.DataFrame,
    *,
    cities: pd.Series | None = None,
) -> pd.DataFrame:
    """Attach rollup ids to a series of household names, preferring a city tiebreak.

    ``households`` needs ``id`` and ``name`` (as ``households_summary`` provides). With
    ``cities`` (aligned to ``names``), a normalized name+city match wins over a name-only
    match — 577 households share a name with another household, so city disambiguates
    Judy's exports. One row per input: ``id`` (string, NA when unmatched) and ``match``
    ("name+city" | "name" | "unmatched"). Unmatched names are surfaced rather than
    fuzzy-guessed — the mailing list must not silently absorb a near-name.

    The result is index-aligned to the input, so it can be assigned onto the caller's
    frame regardless of that frame's index.
    """
    lookup = (
        pd.DataFrame(
            {
                "key": norm_name(households["name"]),
                "id": households["id"],
                **({"city": norm_name(households["city"])} if "city" in households else {}),
            }
        )
        .dropna(subset=["key"])
        .drop_duplicates(subset=["key"] if "city" not in households else ["key", "city"],
                         keep="first")
    )
    out = pd.DataFrame({"key": norm_name(names)})
    out["id"] = pd.Series(pd.NA, index=out.index, dtype="string")
    out["match"] = pd.Series(pd.NA, index=out.index, dtype="object")

    # tier 1: name+city (NaN cities never enter this tier — pandas merges NA==NA, which
    # would multiply blank-city rows against every blank-city household)
    if cities is not None and "city" in lookup.columns:
        tier = lookup.dropna(subset=["city"])
        wants = norm_name(cities).notna()
        hit = (
            pd.DataFrame({"key": out.loc[wants, "key"], "city": norm_name(cities)[wants]})
            .merge(tier, on=["key", "city"], how="left")
        )
        out.loc[wants, "id"] = hit["id"].values
        out.loc[wants, "match"] = np.where(hit["id"].notna().values, "name+city", pd.NA)

    # tier 2: name-only for rows the city tier missed (wrong/blank city, or no city given)
    still = out["id"].isna()
    if still.any():
        name_only = (
            pd.DataFrame({"key": norm_name(households["name"]), "id": households["id"]})
            .dropna(subset=["key"])
            .drop_duplicates(subset=["key"], keep="first")
        )
        fallback = out.loc[still, ["key"]].merge(name_only, on="key", how="left")
        out.loc[still, "id"] = fallback["id"].values
        out.loc[still, "match"] = np.where(fallback["id"].notna().values, "name", "unmatched")
    out["id"] = out["id"].astype("string")
    out.index = names.index  # keep the caller's index for safe .assign alignment
    return out[["id", "match"]]


# -- Judy's donor export --------------------------------------------------------
# workbook header -> canonical column
DONOR3_COLUMNS = {
    "Ambassador": "steward_raw",
    "notes": "note_hand",
    "Household Name/Account Name": "household_name",
    "Primary Contact First Name": "primary_first_name",
    "Primary Contact Last Name": "primary_last_name",
    "Household Salutation / Preferred Name": "salutation",
    "Primary Contact Email": "email",
    "Primary Contact Phone": "phone",
    "Address Line 1": "address_line1",
    "Address Line 2": "address_line2",
    "City": "city",
    "State/Province": "state_province",
    "Zip Code": "zip_code",
    # Judy's fiscal-year totals: "2023-2024 Fiscal Yr" = FY24 (Jul 2023–Jun 2024), etc.
    "2023-2024 Fiscal Yr": "fy2024_amount",
    "2024-2025 Fiscal Yr": "fy2025_amount",
    "2025-2026 Fiscal Yr": "fy2026_amount",
    "Total 3 Yrs": "total_3yr_amount",
}


def clean_steward(raw: pd.Series) -> pd.Series:
    """Distill the free-text Ambassador column into a one-word steward value.

    The hand entries look like ``"Don - know a little"``, ``"Don"``, ``"Sue"``. Everything
    before ``" - "`` is the person; the full original text is kept by the caller as detail.
    """
    return (
        raw.astype("string")
        .str.split(r"\s+-\s+", n=1)
        .str[0]
        .str.strip()
        .str.lower()
        .mask(lambda s: s.eq("") | s.eq("nan"))
    )


def load_donor3(path: Path | None = None) -> pd.DataFrame:
    """Cleaned donor-export rows: household name, steward + note, and Judy's FY amounts."""
    src = Path(path) if path is not None else _ext_path(DONOR3_FILENAME)
    df = pd.read_excel(src, sheet_name="export")
    missing = set(DONOR3_COLUMNS) - set(df.columns)
    if missing:
        raise ValueError(f"{src.name} is missing expected column(s) {sorted(missing)}")
    out = df[list(DONOR3_COLUMNS)].rename(columns=DONOR3_COLUMNS)
    out = out[out["household_name"].notna()].reset_index(drop=True)
    out["excel_row"] = np.arange(len(out)) + 2  # header is Excel row 1
    for c in ("fy2024_amount", "fy2025_amount", "fy2026_amount", "total_3yr_amount"):
        out[c] = pd.to_numeric(out[c], errors="coerce")
    out["steward"] = clean_steward(out["steward_raw"])
    return out


# -- new-accounts export --------------------------------------------------------
NEW_ACCOUNTS_HEADER_ROW = 3  # 0-based row index of the header (workbook row 4)

NEW_ACCOUNTS_COLUMNS = {
    "Household Name/Account Name": "household_name",
    "Primary Contact First Name": "primary_first_name",
    "Primary Contact Last Name": "primary_last_name",
    "Household Salutation / Preferred Name": "salutation",
    "Primary Contact Email": "email",
    "Primary Contact Phone": "phone",
    "Address Line 1": "address_line1",
    "Address Line 2": "address_line2",
    "City": "city",
    "State/Province": "state_province",
    "Zip Code": "zip_code",
    "HH/Acct. All Donation Amount": "lifetime_amount",
    "2026 HH/Acct. Donation Total": "y2026_amount",
    "2025 HH/Acct. Donation Total": "y2025_amount",
    "HH/Acct. All Registration Amount": "lifetime_registration_amount",
}


def load_new_accounts(path: Path | None = None) -> pd.DataFrame:
    """Cleaned new-account rows (FY25–FY26 entrants) with hand notes from the right edge."""
    src = Path(path) if path is not None else _ext_path(NEW_ACCOUNTS_FILENAME)
    df = pd.read_excel(src, sheet_name=0, header=NEW_ACCOUNTS_HEADER_ROW)
    missing = set(NEW_ACCOUNTS_COLUMNS) - set(df.columns)
    if missing:
        raise ValueError(f"{src.name} is missing expected column(s) {sorted(missing)}")
    out = df[list(NEW_ACCOUNTS_COLUMNS)].rename(columns=NEW_ACCOUNTS_COLUMNS)

    # hand notes sit in headerless columns right of the named block (currently only the
    # last one is used, but accept any trailing Unnamed column so future notes aren't lost)
    trailing = [c for c in df.columns if str(c).startswith("Unnamed")]
    if trailing:
        notes = df[trailing].apply(
            lambda row: "; ".join(str(v) for v in row if pd.notna(v)), axis=1
        )
        out["note_hand"] = notes.mask(notes == "").astype("string")
    else:
        out["note_hand"] = pd.Series(pd.NA, index=out.index, dtype="string")

    out = out[out["household_name"].notna()].reset_index(drop=True)
    out["excel_row"] = np.arange(len(out)) + NEW_ACCOUNTS_HEADER_ROW + 2
    for c in ("lifetime_amount", "y2026_amount", "y2025_amount", "lifetime_registration_amount"):
        out[c] = pd.to_numeric(out[c], errors="coerce")
    return out


# -- Don's edited donor workbook -------------------------------------------------
def load_djb_workbook(path: Path | None = None) -> dict[str, pd.DataFrame]:
    """The two hand-marked selections from Don's copy of the donor workbook.

    Returns ``{"silent_selected": ..., "appeal_responded": ...}``:

    - ``silent_selected`` — rows of ``silent-1000plus`` whose household name is *bold*
      (Don's keep-list of non-responding $1k+ households), with the sheet's ``note`` and
      his ``action??`` column combined into one note. Needs openpyxl in normal (not
      read-only) mode because pandas discards cell fonts.
    - ``appeal_responded`` — ``donors`` rows with appealed=TRUE and responded=TRUE (last
      year's appeal responders), note field included.

    Both are keyed by household name exactly as the generated workbook wrote them, so they
    match the rollup without normalization in practice; ``match_households`` still applies.
    """
    src = Path(path) if path is not None else _ext_path(DJB_WORKBOOK_FILENAME)
    wb = openpyxl.load_workbook(src)  # fonts require the standard loader

    ws = wb["silent-1000plus"]
    headers = [c.value for c in ws[1]]
    col = {name: i for i, name in enumerate(headers)}
    selected = []
    for row in ws.iter_rows(min_row=2):
        name = row[col["household"]].value
        if name is None:
            continue
        bold = bool(row[col["household"]].font and row[col["household"]].font.bold)
        if not bold:
            continue
        note = row[col["note"]].value if "note" in col else None
        action = row[col["action??"]].value if "action??" in col else None
        combined = "; ".join(str(v).strip() for v in (note, action) if v)
        selected.append(
            {"household_name": str(name).strip(), "note_hand": combined or pd.NA}
        )
    silent = pd.DataFrame(selected, columns=["household_name", "note_hand"])

    donors = pd.read_excel(src, sheet_name="donors")
    responded = donors[donors["appealed"] == True]  # noqa: E712 — Excel booleans
    responded = responded[responded["responded"] == True]  # noqa: E712
    appealed = responded.rename(
        columns={"household": "household_name", "note": "note_hand"}
    )[["household_name", "note_hand"]].reset_index(drop=True)

    return {"silent_selected": silent, "appeal_responded": appealed}
