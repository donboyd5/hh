"""Mailing-list input loaders (synthetic fixtures; the real workbooks are local-only PII)."""
import openpyxl
import pandas as pd

from hh.external.mailing import (
    clean_steward,
    load_djb_workbook,
    load_donor3,
    load_new_accounts,
    match_households,
    norm_name,
)


def test_norm_name_collapses_internal_whitespace():
    s = pd.Series(["Andrea  Strebel", " Kathy & Hugh Roome ", None, "A&B"])
    out = norm_name(s)
    assert out[0] == "andrea strebel"
    assert out[1] == "kathy & hugh roome"
    assert out[2] is pd.NA
    assert out[3] == "a&b"


def test_match_households_exact_and_unmatched():
    households = pd.DataFrame(
        {"id": ["10", "20", "20"], "name": ["Kathy & Hugh Roome", "Ann Lee", "Ann Lee"]}
    )
    m = match_households(pd.Series(["Kathy  & Hugh Roome", "Nobody"]), households)
    assert m["id"].tolist() == ["10", pd.NA]
    assert m["match"].tolist() == ["name", "unmatched"]


def test_match_households_city_tiebreak_beats_name_only():
    households = pd.DataFrame(
        {
            "id": ["10", "20"],
            "name": ["Ann Lee", "Ann Lee"],
            "city": ["Cambridge", "Salem"],
        }
    )
    m = match_households(
        pd.Series(["Ann Lee", "Ann Lee", "Ann Lee"]),
        households,
        cities=pd.Series(["salem", "Buskirk", None]),
    )
    assert m["id"].tolist() == ["20", "10", "10"]
    assert m["match"].tolist() == ["name+city", "name", "name"]  # blank city falls back


def test_clean_steward_takes_person_before_detail():
    raw = pd.Series(
        ["Don - know a little", "Sue", "Don", "?? board member", None, "Don - know"]
    )
    out = clean_steward(raw)
    assert out.tolist() == ["don", "sue", "don", "?? board member", pd.NA, "don"]


def _write_donor3(path):
    df = pd.DataFrame(
        {
            "Ambassador": ["Don - know a little", None],
            "notes": ["friend of Alix", None],
            "Household Name/Account Name": ["Kathy & Hugh Roome", None],
            "Primary Contact First Name": ["Kathy", None],
            "Primary Contact Last Name": ["Roome", None],
            "Household Salutation / Preferred Name": ["Kathy and Hugh", None],
            "Primary Contact Email": ["k@x.org", None],
            "Primary Contact Phone": ["518-555-0100", None],
            "Address Line 1": ["1 Main St", None],
            "Address Line 2": [None, None],
            "City": ["Cambridge", None],
            "State/Province": ["NY", None],
            "Zip Code": ["12816", None],
            "2023-2024 Fiscal Yr": [100, None],
            "2024-2025 Fiscal Yr": [250, None],
            "2025-2026 Fiscal Yr": [0, None],
            "Total 3 Yrs": [350, None],
            "sort": [1, 2],
        }
    )
    with pd.ExcelWriter(path) as xw:
        df.to_excel(xw, sheet_name="export", index=False)


def test_load_donor3_renames_and_drops_blank_rows(tmp_path):
    p = tmp_path / "d3.xlsx"
    _write_donor3(p)
    out = load_donor3(p)
    assert len(out) == 1
    row = out.iloc[0]
    assert row["household_name"] == "Kathy & Hugh Roome"
    assert row["steward"] == "don"
    assert row["steward_raw"] == "Don - know a little"
    assert row["note_hand"] == "friend of Alix"
    assert row["fy2024_amount"] == 100 and row["fy2026_amount"] == 0
    assert row["excel_row"] == 2  # first data row


def _write_new_accounts(path):
    # header on workbook row 4 (index 3), two headerless note columns at the right edge
    cols = [
        "Primary Contact First Name",
        "Primary Contact Last Name",
        "Household Name/Account Name",
        "Household Salutation / Preferred Name",
        "Address Line 1",
        "Address Line 2",
        "City",
        "State/Province",
        "Zip Code",
        "Primary Contact Email",
        "Primary Contact Phone",
        "HH/Acct. Last Donation Date",
        "HH/Acct. All Donation Amount",
        "HH/Acct. Largest Donation Amount",
        "2026 HH/Acct. Donation Total",
        "2025 HH/Acct. Donation Total",
        "HH/Acct. Last Donation Campaign",
        "HH/Acct. All Registration Amount",
        "HH/Acct. Last Registration Date",
    ]
    data = pd.DataFrame(
        [
            ["Robyn", "Schanzlin", "Robyn & Michael Schanzlin", "Robyn and Michael",
             "16 E Main St", "", "Cambridge", "NY", "12816", "r@x.org", "518-555-0101",
             None, 0, 0, 0, 0, None, 1427, "2026-05-11"],
            ["Leah", "Tinkham", "Leah Tinkham", "Leah", "2 Elm St", "", "Salem", "NY",
             "12865", "l@x.org", None, None, None, None, None, None, None, None, None],
        ],
        columns=cols,
    )
    with pd.ExcelWriter(path) as xw:
        pd.DataFrame({"junk": ["genlly came to event"]}).to_excel(
            xw, sheet_name="New Accounts 25-26", index=False, startrow=0
        )
        data.to_excel(
            xw, sheet_name="New Accounts 25-26", index=False, startrow=3, header=True
        )
        notes = pd.DataFrame([["Robyn King", "came to gala"], [None, None]])
        notes.to_excel(
            xw,
            sheet_name="New Accounts 25-26",
            index=False,
            startrow=4,
            startcol=19,  # headerless right-edge columns, past the named block
            header=False,
        )


def test_load_new_accounts_reads_header_row_4_and_notes(tmp_path):
    p = tmp_path / "na.xlsx"
    _write_new_accounts(p)
    out = load_new_accounts(p)
    assert len(out) == 2
    row = out.iloc[0]
    assert row["household_name"] == "Robyn & Michael Schanzlin"
    assert row["note_hand"] == "Robyn King; came to gala"
    assert row["excel_row"] == 3 + 2  # header row index 3 -> first data row is Excel row 5
    assert out.iloc[1]["note_hand"] is pd.NA


def _write_djb(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "silent-1000plus"
    ws.append(["household", "town", "note", "action??"])
    rows = [
        ("Hannah Stevens", "Cambridge", "was on David hiring committee", "sustainer; Don"),
        ("Plain Person", "Salem", None, None),  # not bold -> dropped
        ("Bruce Merrill", "Greenwich", "Judy met at Gala", "letter"),
    ]
    for name, town, note, action in rows:
        ws.append([name, town, note, action])
        ws.cell(row=ws.max_row, column=1).font = openpyxl.styles.Font(
            bold=name != "Plain Person"
        )
    ws2 = wb.create_sheet("donors")
    ws2.append(["household", "appealed", "responded", "note"])
    for household, appealed, responded, note in [
        ("A Donor", True, True, "gave $500"),
        ("B Donor", True, False, None),
        ("C Donor", False, True, "never appealed"),
    ]:
        ws2.append([household, appealed, responded, note])
    wb.save(path)


def test_load_djb_workbook_bolds_notes_and_responded(tmp_path):
    p = tmp_path / "djb.xlsx"
    _write_djb(p)
    out = load_djb_workbook(p)
    silent = out["silent_selected"]
    assert silent["household_name"].tolist() == ["Hannah Stevens", "Bruce Merrill"]
    assert silent.iloc[0]["note_hand"] == (
        "was on David hiring committee; sustainer; Don"
    )
    responded = out["appeal_responded"]
    assert responded["household_name"].tolist() == ["A Donor"]
    assert responded.iloc[0]["note_hand"] == "gave $500"
